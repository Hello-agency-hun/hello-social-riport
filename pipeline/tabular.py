"""Egységes táblázat-beolvasás, a fájlnévtől és kiterjesztéstől függetlenül."""

from datetime import date, datetime
from pathlib import Path
import zipfile

from pipeline.textio import read_csv_header, read_csv_rows

OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _records(rows) -> list[dict[str, str]]:
    materialized = list(rows)
    if not materialized:
        return []
    header = [_text(value).strip().lstrip("\ufeff") for value in materialized[0]]
    if not any(header):
        return []
    return [
        {column: _text(value) for column, value in zip(header, row)}
        for row in materialized[1:]
        if any(_text(value).strip() for value in row)
    ]


def _xlsx_rows(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            records = _records(sheet.iter_rows(values_only=True))
            if records:
                return records
        return []
    finally:
        workbook.close()


def _xls_rows(path: Path) -> list[dict[str, str]]:
    import xlrd

    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        for sheet in workbook.sheets():
            records = _records(sheet.row_values(index) for index in range(sheet.nrows))
            if records:
                return records
        return []
    finally:
        workbook.release_resources()


def table_format(path: Path) -> str:
    """A tényleges táblázatformátum; a kiterjesztés csak támpont."""
    path = Path(path)
    with path.open("rb") as stream:
        head = stream.read(8)
    if head.startswith(OLE_MAGIC):
        return "xls"
    if zipfile.is_zipfile(path):
        return "xlsx"
    return "csv"


def read_table_rows(path: Path) -> list[dict[str, str]]:
    """CSV, XLSX vagy XLS első nem üres munkalapja szótársorokként."""
    actual = table_format(path)
    if actual == "xlsx":
        return _xlsx_rows(path)
    if actual == "xls":
        return _xls_rows(path)
    return read_csv_rows(path)


def read_table_header(path: Path) -> list[str]:
    path = Path(path)
    actual = table_format(path)
    if actual == "csv":
        return read_csv_header(path)
    rows = read_table_rows(path)
    if rows:
        return list(rows[0])

    # A fejléc önmagában is elég a forrástípus felismeréséhez. Az üres export
    # később a parserben kap pontos, emberi hibaüzenetet.
    if actual == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                row = next(sheet.iter_rows(values_only=True), ())
                header = [_text(value).strip().lstrip("\ufeff") for value in row]
                if any(header):
                    return header
        finally:
            workbook.close()
    elif actual == "xls":
        import xlrd

        workbook = xlrd.open_workbook(str(path), on_demand=True)
        try:
            for sheet in workbook.sheets():
                if sheet.nrows:
                    header = [_text(value).strip().lstrip("\ufeff") for value in sheet.row_values(0)]
                    if any(header):
                        return header
        finally:
            workbook.release_resources()
    return []
