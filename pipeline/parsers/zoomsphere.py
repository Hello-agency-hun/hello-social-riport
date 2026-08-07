import re
import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from pipeline.errors import MissingColumnError
from pipeline.schema import ContentItem, ParsedSource

REQUIRED_HEADERS = {"PostType", "FacebookPostIDs", "InstagramPostIDs"}


def looks_like_zoomsphere(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as zf:
            sheet = zf.read("xl/worksheets/sheet1.xml").decode("utf-8", "replace")
    except (zipfile.BadZipFile, KeyError):
        return False
    return all(header in sheet for header in REQUIRED_HEADERS)


# A médiaoszlopok sorrendje számít: a riport a lista első elemét mutatja
# thumbnailként. Videónál és reelnél a nyers fájl mp4, amiből nem lesz kép —
# ezért a poszter-kép (`*VideoThumbnail`) előrébb áll, mint a videó URL-je.
CHANNEL_COLUMNS = {
    "facebook": {
        "message": "FacebookMessage",
        "source": "FacebookSources",
        "ids": "FacebookPostIDs",
        "permalink": "FacebookPublicPermalinks",
        "images": [
            "FacebookImages",
            "FacebookFileUrl",
            "FacebookVideoThumbnail",
            "FacebookFileThumbnail",
            "FacebookVideoUrl",
        ],
    },
    "instagram": {
        "message": "InstagramMessage",
        "source": "InstagramSources",
        "ids": "InstagramPostIDs",
        "permalink": "InstagramPublicPermalinks",
        "images": [
            "InstagramImages",
            "InstagramFileUrl",
            "InstagramVideoThumbnail",
            "InstagramFileThumbnail",
            "InstagramVideoUrl",
        ],
    },
}


def _clean(value: str) -> str:
    """A ZoomSphere minden azonosító mögé odaírja a fiók nevét zárójelben."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", (value or "").strip())


def _post_id(raw: str) -> str:
    """FB: `oldal_poszt` → `poszt`. IG: változatlan."""
    cleaned = _clean(raw)
    return cleaned.split("_")[-1] if cleaned else ""


def _urls(value: str) -> list[str]:
    return [u.strip() for u in (value or "").split(",") if u.strip().startswith("http")]


DATETIME_FORMAT = "%d.%m.%Y - %I:%M %p"


def _parse_datetime(value: str) -> date:
    """`01.07.2026 - 11:00 AM` → date(2026, 7, 1)

    Ha a ZoomSphere megváltoztatja a formátumot — vagy valaki megnyitja és
    menti a fájlt Excelben, ami átírja a dátumokat —, itt nyers `ValueError`
    jött, `_strptime` stack trace-szel. Abból a menedzser nem tudhatja, hogy a
    fájllal van baj, nem a programmal.
    """
    try:
        return datetime.strptime((value or "").strip(), DATETIME_FORMAT).date()
    except ValueError as error:
        raise MissingColumnError(
            f"a ZoomSphere export `Datetime` oszlopában értelmezhetetlen dátum: "
            f"{value!r}. Várt formátum: `01.07.2026 - 11:00 AM`.\n"
            "Gyakori ok: a fájlt megnyitották Excelben és mentették — az átírja "
            "a dátumokat. Töltsd le újra, és ne nyisd meg."
        ) from error


def parse(path) -> ParsedSource:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    index = {name: position for position, name in enumerate(header)}

    items: list[ContentItem] = []
    page_name = ""

    for row in rows:
        cells = ["" if cell is None else str(cell) for cell in row]
        cells += [""] * (len(header) - len(cells))

        def column(name: str) -> str:
            return cells[index[name]] if name in index else ""

        if not column("Datetime").strip():
            continue

        item = ContentItem(
            published=_parse_datetime(column("Datetime")),
            post_type=column("PostType").strip(),
        )

        for channel, columns in CHANNEL_COLUMNS.items():
            if not column(columns["source"]).strip():
                continue
            page_name = page_name or _clean(column(columns["source"]))
            item.captions[channel] = column(columns["message"]).strip()
            item.post_ids[channel] = _post_id(column(columns["ids"]))
            item.permalinks[channel] = _clean(column(columns["permalink"]))
            creatives: list[str] = []
            for image_column in columns["images"]:
                creatives += _urls(column(image_column))
            item.creatives[channel] = creatives

        items.append(item)

    workbook.close()
    dates = sorted(item.published for item in items)
    return ParsedSource(
        kind="zoomsphere",
        period=(dates[0], dates[-1]),
        client_hints={"page_name": page_name},
        payload=items,
    )
